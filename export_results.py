import os
import sys
from numpy import column_stack
import pandas as pd
from openpyxl import load_workbook

def _default_output_path(filename: str = "results.xlsx") -> str:
    # Get the default output path
	if getattr(sys, "frozen", False): base_dir = os.path.dirname(sys.executable)
	else: base_dir = os.path.dirname(os.path.abspath(__file__))
	return os.path.join(base_dir, filename)

def save_excel(rows, output_path: str = None, sheet_name: str = None) -> str:
    # Save data to the specified sheet
    if sheet_name == "DB":
        columns = ["title", "date", "exc_amount", "exc_shares", "exc_price", "listing_date"]
    elif sheet_name == "EX":
        columns = ["title", "date", "prv_prc", "cur_prc"]

    # Create a dataframe from the rows
    df = pd.DataFrame(rows)
    available = [c for c in columns if c in df.columns]
    df = df[available]
    
    if not output_path: output_path = _default_output_path()
    output_dir = os.path.dirname(output_path) or "."
    os.makedirs(output_dir, exist_ok=True)

    # Load the workbook
    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    # Write header
    header_present = False
    if ws.max_row >= 1:
        for col_idx, col_name in enumerate(available, start=1):
            if ws.cell(row=1, column=col_idx).value:
                header_present = True
                break
    if not header_present:
        for col_idx, col_name in enumerate(available, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    start_row = ws.max_row + 1
    for offset, (_, row) in enumerate(df.iterrows(), start=0):
        for col_idx, col_name in enumerate(available, start=1):
            ws.cell(row=start_row + offset, column=col_idx, value=row.get(col_name))

    wb.save(output_path)
    return os.path.abspath(output_path)

def read_list_titles(output_path: str = None) -> list:
    # Read target companies from the LIST sheet
	if not output_path: output_path = _default_output_path()
	wb = load_workbook(output_path, read_only=True, data_only=True)
	ws = wb["LIST"]
	values = []
	for row_idx in range(2, ws.max_row + 1):
		cell_value_a = ws.cell(row=row_idx, column=1).value
		cell_value_b = ws.cell(row=row_idx, column=2).value
		text_a = str(cell_value_a).strip()
		text_b = str(cell_value_b).strip()
		values.append([text_a, text_b])
	wb.close()
	return values

def clear_excel(output_path: str = None, sheet_name: str = None) -> str:
    # Clear the specified sheet
	if not output_path: output_path = _default_output_path()
	wb = load_workbook(output_path)
	ws = wb[sheet_name]
	ws.delete_rows(1, ws.max_row)
	wb.save(output_path)
	return os.path.abspath(output_path)
